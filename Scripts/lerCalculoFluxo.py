import openpyxl
from openpyxl import load_workbook
import numpy as np
import pandas as pd

def lerCF(pathCF, numeroCCI):

    pathExcel = pathCF
    numCCI = numeroCCI

    #Carregar Planilha
    wb = load_workbook(filename=pathExcel, 
                   read_only=True)
    try:
        ws = wb['PRICE_CORR']
    except:
        ws = wb['SAC_CORR']

    #Percorrer Coluna B e transformar em DF
    data_rows = []
    for row in ws['A18':'A500']:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    #Achar ultima linha e criar DF
    df = pd.DataFrame(data_rows)
    df.replace('',np.nan, inplace=True)
    df = df.dropna()
    lastRow = len(df) + 17
    df.columns = ['numParcela']
    try:
        df['numParcela'] = df['numParcela'].replace("I", 0)
    except:
        print('erro')
    df['numParcela'] = df['numParcela'].astype(int)

    #Percorrer Coluna B e transformar em DF
    date_rows = []
    
    print(f'\nLastRow: {lastRow}\n')
    
    for row in ws['B18':f'B{lastRow}']:
        date_cols = []
        for cell in row:
            date_cols.append(cell.value)
        date_rows.append(date_cols)
    df["data_vencto"] = pd.DataFrame(date_rows)

    #Percorrer Coluna C e transformar em DF
    date_rows = []
    for row in ws['C18':f'C{lastRow}']:
        date_cols = []
        for cell in row:
            date_cols.append(cell.value)
        date_rows.append(date_cols)
    df["amortizacao"] = pd.DataFrame(date_rows)

    #Percorrer Coluna D e transformar em DF
    date_rows = []
    for row in ws['D18':f'D{lastRow}']:
        date_cols = []
        for cell in row:
            date_cols.append(cell.value)
        date_rows.append(date_cols)
    df["juros"] = pd.DataFrame(date_rows)

    
    #Atribuir Colunas DF
    df['juros'] = df['juros'].replace(np.nan, 0)
    df['valor_principal'] = round(df['amortizacao'],2)
    df['valor_futuro'] = round(df['amortizacao']+df['juros'],2)

    #Valor CCI
    df['codigo_integracao'] = numCCI

    #Valores Padrões    
    df['id_tipo_parcela'] = "2"
    df['Aplica_Correcao'] = "S"
    df['Aplica_Juros'] = "S"
    df['periodicidade'] = "1" 
    
    print(f'\nDF Antigo:\n{df}\n')

    #Gerar Novo DF
    df = df[['codigo_integracao', 'data_vencto', 'id_tipo_parcela', 'numParcela', 'valor_principal', 'valor_futuro', 'Aplica_Correcao', 'Aplica_Juros', 'periodicidade']]
    df['data_vencto'] = pd.to_datetime(df['data_vencto'], format='%d/%m/%y').dt.strftime('%d/%m/%Y')

    #Tratar DF
    df_remove = df.loc[(df['numParcela'] == 0)]
    df = df.drop(df_remove.index)

    print(f'\nDF Novo:\n{df}\n')

    if df['numParcela'].iloc[0] == 2:
        df['numParcela'] = df['numParcela'] - 1

    df['numParcela'] = np.where(df['valor_principal'] >0, df['numParcela'].apply(str)+"/"+str(df['numParcela'].iloc[-1]), 0)
    df = df.rename(columns={'numParcela': 'numero_parcela'})
    
    #Gerar Json
    jsonValue = df.to_json(orient='records').replace("\\/", str('/'))

    return jsonValue

###### teste #######
# pathExcel = r"G:\Drives compartilhados\Pontte Crédito\0_HOME EQUITY\0_Analises\MARIA SONIA PAULO DA SILVA ID 586505100\KIT QI\Cálculo_Fluxo.xlsx"
# numCCI = 123456

# from pprint import pprint
# pprint(lerCF(pathExcel, numCCI))