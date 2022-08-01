import openpyxl as xl
from time import sleep
from datetime import datetime
import warnings

#Remover Avisos
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def lerSimulador(path):
    wb = xl.load_workbook(path, data_only=True)
    ws3 = wb['Simulador PDF']

    #Variáveis do Quadro Simulação
    amortizacao = None
    indexador = None
    prazo = None
    carencia = None
    pularParcela = None
    diaVencimento = None
    taxaAM = None
    CET = None

    #Variáveis do do Quadro Resultados
    LTV = None
    IOF = None
    TAC = None
    registro = None
    valorBruto = None
    valorLiquido = None
    valorAvaliacao = None
    primeiraParcela = None

    #Variáveis do Quadro de Parcela
    auxPar = False
    dataPriParcela = None
    dataDesembolco = None
    priParcelaAmortizao = None
    priParcelaJuros = None
    dataUltimaParcela = 0
    auxSaldo = False

    #Extração de Conteúdo da Planilha
    for col in ws3.iter_cols(min_col=1,max_col=15, min_row = 1, max_row=100):
        for cell in col:
            
            #Ler Quadro Simulação
            if cell.value == 'Simulação':
                #Definir variáveis auxiliares de linha e coluna
                rowAux = cell.row + 1
                columnAux = cell.column + 3

                #Extração da tabela simulação
                amortizacao = ws3.cell(row=rowAux, column=columnAux).value
                indexador = ws3.cell(row=rowAux+1, column=columnAux).value
                prazo = ws3.cell(row=rowAux+2, column=columnAux).value
                carencia = ws3.cell(row=rowAux+3, column=columnAux).value
                pularParcela = ws3.cell(row=rowAux+4, column=columnAux).value
                diaVencimento = ws3.cell(row=rowAux+5, column=columnAux).value
                taxaAM = ws3.cell(row=rowAux+6, column=columnAux).value
                CET = ws3.cell(row=rowAux+7, column=columnAux).value
            
            #Ler Quadro Resultados
            elif cell.value == 'Resultados':
                #Definir variáveis auxiliares de linha e coluna
                rowAux = cell.row + 1
                columnAux = cell.column + 3

                #Extralção da tabela Resultados
                LTV = ws3.cell(row=rowAux, column=columnAux).value
                IOF = ws3.cell(row=rowAux+1, column=columnAux).value
                TAC = ws3.cell(row=rowAux+2, column=columnAux).value
                registro = ws3.cell(row=rowAux+3, column=columnAux).value
                valorBruto = ws3.cell(row=rowAux+4, column=columnAux).value
                valorLiquido = ws3.cell(row=rowAux+5, column=columnAux).value
                valorAvaliacao = ws3.cell(row=rowAux+6, column=columnAux).value
                primeiraParcela = ws3.cell(row=rowAux+7, column=columnAux).value
            
            #Ler Quadro de Parcelas
            elif cell.value == 'Data':
                rowAux = cell.row + 1
                columnAux = cell.column + 1
                dataDesembolco = ws3.cell(row=rowAux, column=cell.column).value
                for i in range(1,10):
                    parcelaMensal = ws3.cell(row=rowAux+i, column=columnAux).value
                    if parcelaMensal > 0 and auxPar == False:
                        priParcelaAmortizao = ws3.cell(row=rowAux+i, column=columnAux+1).value
                        priParcelaJuros = ws3.cell(row=rowAux+i, column=columnAux+2).value
                        dataPriParcela = ws3.cell(row=rowAux+i, column=columnAux-1).value
                        #print(f"Parcela do mês: {parcelaMensal} \nAmortização: {priParcelaAmortizao}\nJuros: {priParcelaJuros}\nAmortização+Juros: {str(priParcelaJuros+priParcelaAmortizao)}")
                        auxPar = True
            
            #Ler Coluna Saldo Devedor
            elif cell.value == 'Saldo Devedor':
                for i in range(1,400):
                    saldoDevedor = ws3.cell(row=rowAux+i, column=cell.column).value
                    try:
                        if saldoDevedor <= 10 and auxSaldo == False:
                            dataUltimaParcela = ws3.cell(row=rowAux+i, column=cell.column-4).value
                            auxSaldo = True
                    except:
                        pass

    #Calculos de Valores Auxiliares
    amortizacaoEJuros = priParcelaAmortizao + priParcelaJuros
    prestacoes = prazo - carencia

    #Pegar nome do Cliente
    nomeCliente = ws3.cell(row=1, column=7).value

    #Fechar Planilha
    wb.close()

    #Criar Chave Valor com Resultados de Saída
    keyValue = {'valorTotal':round(valorBruto,2),'tabela': amortizacao,'registro': round(registro,2),'taxaAM':taxaAM,'valorLiquido': round(valorLiquido,2),'prazoMes': prazo,
                'valorPrimeiraParcelaComEncargos':round(primeiraParcela,2), 'valorImóvel':round(valorAvaliacao,2),'prazoContrato': prestacoes,'ultimaParcela':dataUltimaParcela.strftime("%d/%m/%Y"),
                'dataContrato': dataDesembolco.strftime("%d/%m/%Y"), 'valorPrimeiraParcela': round(amortizacaoEJuros,2), 'primeiraParcela':dataPriParcela.strftime("%d/%m/%Y"),'indice': indexador
                }
    
    #Retornar Chave e Valor
    return keyValue
